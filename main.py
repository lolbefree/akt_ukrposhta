import math
import os
import sys
import openpyxl
import pyodbc
from openpyxl.styles import Alignment, Font
import decimal
import sql_querys
import not_for_git
from babel.dates import format_datetime
from openpyxl.styles.borders import Border, Side

col_names = ["№ з/п", "№ по каталогу", "Найменування матеріалів, марка", "Код \nУКТ ЗЕД", "Од. виміру", "Кількість",
             "Ціна за одиницю\n (без ПДВ). грн.", "Вартість (без ПДВ), грн."]


class ActPost:
    def __init__(self, gsalid):
        self.gsalid = gsalid
        self.file_path = os.path.abspath(os.path.dirname(__file__))
        self.wb = openpyxl.load_workbook(f"{self.file_path}\\template.xlsx")
        self.ws1 = self.wb[self.wb.sheetnames[0]]
        self.server = not_for_git.db_server
        self.database = not_for_git.db_name
        self.username = not_for_git.db_user
        self.password = not_for_git.db_pw
        self.driver = '{SQL Server}'  # Driver you need to connect to the database
        self.port = '1433'
        self.cnn = pyodbc.connect(
            'DRIVER=' + self.driver + ';PORT=port;SERVER=' + self.server + ';PORT=1443;DATABASE=' + self.database +
            ';UID=' + self.username +
            ';PWD=' + self.password)
        self.cursor = self.cnn.cursor()

        self.thin_border = Border(left=Side(style='medium'),
                                  right=Side(style='medium'),
                                  top=Side(style='medium'),
                                  bottom=Side(style='medium'))
        self.center = Alignment(horizontal='center')
        self.save_file_name = ""
        self.central_table = list(self.cursor.execute(sql_querys.central_table(self.gsalid)))
        self.price_cnt = int()
        self.price_footer = int()
        self.price_hour = 560.00
        self.central_table_sum = decimal.Decimal()

    def round_half_up(self, n, decimals=0):
        multiplier = 10 ** decimal.Decimal(decimals)

        return math.floor(n * multiplier + decimal.Decimal(0.5)) / multiplier

    def get_headed(self):
        header = list(self.cursor.execute(sql_querys.get_header(self.gsalid)))[0]
        self.save_file_name = f"{header[0]}{header[1]}".replace("/", "-")
        self.ws1["F1"] = f"№ {header[0]}{header[1]}"
        self.ws1["g2"] = format_datetime(header[2], "dd.MM.Y", locale='uk_UA') if header[2] else " "
        self.ws1["b14"] = header[4]
        self.ws1["g11"] = header[-1]
        self.ws1["d14"] = header[5]
        self.ws1["e14"] = header[6]
        self.ws1["f14"] = format_datetime(header[7], "dd.MM.Y", locale='uk_UA') if header[7] else " "
        self.ws1["g14"] = header[8]
        self.ws1["H14"] = format_datetime(header[9], "dd.MM.Y", locale='uk_UA') if header[9] else " "
        self.ws1["I14"] = format_datetime(header[10], "dd.MM.Y", locale='uk_UA') if header[10] else " "
        self.ws1["I11"] = format_datetime(header[-2], "dd.MM.Y", locale='uk_UA') if header[-2] else " "

    def get_middle(self) -> int:
        start_cell_index_central = 18
        if len(self.central_table) == 0:
            self.central_table = [["", "", "", "", "", "", ]]
            self.price_cnt = 0
        else:
            self.price_cnt = self.central_table[0][-1]

        for num, row in enumerate(self.central_table, 1):
            self.ws1.cell(row=start_cell_index_central, column=2).value = num
            self.ws1.cell(row=start_cell_index_central, column=3).value = self.central_table[num - 1][0]
            self.ws1.cell(row=start_cell_index_central, column=4).value = self.central_table[num - 1][1]

            self.ws1.cell(row=start_cell_index_central, column=5).value = self.central_table[num - 1][2]
            # self.ws1.merge_cells(f"f{start_cell_index_central}:g{start_cell_index_central}")
            self.ws1.cell(row=start_cell_index_central, column=6).value = self.central_table[num - 1][3]
            self.ws1.cell(row=start_cell_index_central, column=7).value = str(self.price_hour).replace(".", ",")
            price_with_pdv = decimal.Decimal(self.price_hour) * decimal.Decimal(self.central_table[num - 1][3])
            self.ws1.cell(row=start_cell_index_central, column=8).value = "{:.2f}".format(price_with_pdv).replace(".",
                                                                                                                  ",")
            self.ws1.cell(row=start_cell_index_central, column=9).value = "{:.2f}".format(price_with_pdv).replace(".",
                                                                                                                  ",")
            self.central_table_sum += price_with_pdv
            start_cell_index_central += 1

        start_cell_index_central = 18
        for num in self.central_table:
            for item in range(2, 10):
                self.ws1.cell(row=start_cell_index_central, column=item).border = self.thin_border
                self.ws1.cell(row=start_cell_index_central, column=item).alignment = self.center
                self.ws1.cell(row=start_cell_index_central, column=item).font = Font(size=10, bold=True)
            start_cell_index_central += 1

        start_cell_index_central = 18
        row = start_cell_index_central + len(self.central_table)
        self.ws1.cell(row=row, column=5).border = self.thin_border
        self.ws1.cell(row=row, column=6).border = self.thin_border
        self.ws1.cell(row=row, column=7).border = self.thin_border
        self.ws1.cell(row=row, column=8).border = self.thin_border
        self.ws1.cell(row=row, column=9).border = self.thin_border
        self.ws1.cell(row=row, column=6).alignment = self.center
        self.ws1.cell(row=row, column=7).alignment = self.center

        self.ws1.cell(row=start_cell_index_central + len(self.central_table),
                      column=5).value = "Всього:"

        self.ws1.cell(row=start_cell_index_central + len(self.central_table),
                      column=5).font = Font(size=10, bold=True)

        try:
            time_count = sum([i[-2] for i in self.central_table])
        except TypeError:
            time_count = ""
        self.ws1.cell(row=start_cell_index_central + len(self.central_table), column=6).value = time_count
        self.ws1.cell(row=start_cell_index_central + len(self.central_table), column=6).font = Font(size=10, bold=True)
        self.ws1.cell(row=start_cell_index_central + len(self.central_table), column=9).value = "{:.2f}".format(
            self.central_table_sum).replace(".", ",")
        self.ws1.cell(row=start_cell_index_central + len(self.central_table), column=9).font = Font(size=10, bold=True)
        self.ws1.cell(row=start_cell_index_central + len(self.central_table), column=9).alignment = self.center

        return len(self.central_table) + start_cell_index_central + 2

    def get_footer(self, end_central_table):

        footer = list(self.cursor.execute(sql_querys.footer(self.gsalid)))

        if footer:
            self.price_footer = sum([i[-1] for i in footer])

        if len(footer) == 0:
            footer = [["", "", "", "", "", "", "", "", "", ]]
            self.price_footer = 0

        self.ws1.merge_cells(f"b{end_central_table}:i{end_central_table}")
        self.ws1[f"b{end_central_table}"].value = "Запасні частини та витратні матеріали Виконавця," \
                                                  " використані для надання Послуг (з/ч)"
        self.ws1[f"b{end_central_table}"].alignment = self.center
        self.ws1[f"b{end_central_table}"].font = Font(size=9, bold=True)
        end_central_table += 1

        letters = ["b", "c", "d", "e", "f", "g", "h", "i"]

        for col, name, num in zip(letters, col_names, range(2, len(letters) + 2)):

            self.ws1.merge_cells(f"{col}{end_central_table}:{col}{end_central_table + 3}")
            # self.ws1[f"{col}{end_central_table}"].alignment.vertical = self.center
            self.ws1[f"{col}{end_central_table}"].alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=True)
            self.ws1[f"{col}{end_central_table}"].font = Font(size=6)
            self.ws1[f"{col}{end_central_table}"].value = name
            for i in range(4):
                self.ws1.cell(row=end_central_table + i, column=num).border = self.thin_border
        end_central_table += 4
        cnt = 1

        def input_data(data) -> None:
            nonlocal end_central_table, cnt
            data = list(data)
            data.insert(0, cnt)
            data = data[0:6] + data[-2:]
            for col_, name_, num_ in zip(letters, data, range(2, len(letters) + 2)):
                self.ws1[f"{col_}{end_central_table}"].alignment = Alignment(horizontal='center', vertical='center',
                                                                             wrap_text=True)
                self.ws1[f"{col_}{end_central_table}"].font = Font(size=10, bold=True)
                self.ws1[f"{col_}{end_central_table}"].value = name_
                self.ws1.cell(row=end_central_table, column=num_).border = self.thin_border

            end_central_table += 1
            cnt += 1

        for row in footer:
            input_data(row)
        self.ws1[f"e{end_central_table}"].value = "Всього:"
        try:
            self.ws1[f"i{end_central_table}"].value = "{:.2f}".format(sum([i[-1] for i in footer])).replace(".", ",") \
                if footer else ""
        except TypeError:
            self.ws1[f"i{end_central_table}"].value = ""

        self.ws1[f"e{end_central_table}"].alignment = self.center
        self.ws1[f"i{end_central_table}"].alignment = self.center
        self.ws1[f"e{end_central_table}"].font = Font(size=10, bold=True)
        self.ws1[f"i{end_central_table}"].font = Font(size=10, bold=True)
        for cell in range(5, 10):
            self.ws1.cell(end_central_table, cell).border = self.thin_border

        end_central_table += 2
        pdv = ((self.price_cnt + self.price_footer) * decimal.Decimal(1.2)) - (
                self.price_cnt + self.price_footer)
        pdv = self.round_half_up(pdv, 2)
        for row in range(1, 4):
            self.ws1.merge_cells(f"e{end_central_table}:g{end_central_table}")
            value_b = "Вартість Послуг (без ПДВ)" if row == 1 else (
                "ПДВ, 20%" if row == 2 else ("Всього з ПДВ" if row == 3 else ""))
            self.ws1[f"e{end_central_table}"].value = value_b
            self.ws1[f"e{end_central_table}"].font = Font(bold=True)
            self.ws1[f"e{end_central_table}"].alignment = Alignment(horizontal='right')
            self.ws1.cell(row=end_central_table, column=9).value = "грн."
            self.ws1.cell(row=end_central_table, column=5).font = Font(bold=True)
            self.ws1.cell(row=end_central_table, column=5).alignment = Alignment(horizontal='right')
            self.ws1.cell(row=end_central_table, column=8).alignment = Alignment(horizontal='right')
            self.ws1.cell(row=end_central_table, column=9).font = Font(bold=True)
            self.ws1.cell(row=end_central_table, column=8).font = Font(bold=True)
            else_price = "{:.2f}".format(self.round_half_up(self.price_cnt + self.price_footer, 2)).replace(".",
                                                                                                            ",") \
                if row == 1 else ("{:.2f}".format(pdv).replace(".", ",") if row == 2 else ("{:.2f}".format(
                self.round_half_up(self.price_cnt + self.price_footer + decimal.Decimal(pdv), 2)).replace(".",
                                                                                                          ",")))
            self.ws1.cell(row=end_central_table, column=8).value = else_price
            end_central_table += 1
        self.ws1.merge_cells(f"b{end_central_table}:i{end_central_table}")
        self.ws1[
            f"b{end_central_table}"].value = "Послуги надані у повному обсязі, претензій до якості послуг у Замовника" \
                                             " немає. Замінені запасні частини отримані у повному обсязі."
        self.ws1[f"b{end_central_table}"].font = Font(size=9, bold=True)
        self.ws1[f"b{end_central_table}"].alignment = Alignment(horizontal='center')
        end_central_table += 2

        for i in range(8):  # футер с восьми натписей
            text_size = 9 if i in (1, 2, 3, 6) else (12 if i == 5 else 5)
            value_b = "Послуги надав:" if i == 1 else ("представник Виконавця" if i == 2 else (
                "Заступник директора по РзКК Алексєєв П.О../_____________/" if i == 3 else (
                    "Керівник Виконавця:" if i == 5 else (
                        "Директор Бєлозьорова Г.О. ./_____________/" if i == 6 else (
                            "М.п.    (підпис)                                        (П.І.Б.)" if i == 7 else "")))))
            value_e = "Послуги прийняв:" if i == 1 else ("представник Замовника	" if i == 2 else (
                "Директор Департаменту транспорту і логістики Менабде Г.Г../_____________/" if i == 3 else (
                    "Керівник Замовника:" if i == 5 else (
                        "Директор Департаменту транспорту і логістики Менабде Г.Г../_____________/" if i == 6 else (
                            "М.п.    (підпис)                                        (П.І.Б.)" if i == 7 else "")))))

            self.ws1.merge_cells(f"b{end_central_table + i}:d{end_central_table + i}")
            self.ws1.merge_cells(f"e{end_central_table + i}:i{end_central_table + i}")
            self.ws1[f"b{end_central_table + i}"].alignment = Alignment(horizontal='center')
            self.ws1[f"e{end_central_table + i}"].alignment = Alignment(horizontal='center')
            self.ws1[f"b{end_central_table + i}"].font = Font(size=text_size, bold=True if i not in (1, 2) else False)
            self.ws1[f"e{end_central_table + i}"].font = Font(size=text_size, bold=True if i not in (1, 2) else False)
            self.ws1[f"e{end_central_table + i}"].value = value_e
            self.ws1[f"b{end_central_table + i}"].value = value_b
            if i == 4:
                end_central_table += 1

            self.ws1.column_dimensions['L'].width = 10

    def create_document(self):
        self.get_headed()
        self.get_footer(self.get_middle())
        self.wb.save(f"C:\\Users\\{os.getlogin()}\\Desktop\\{self.save_file_name}.xlsx")
        os.system(f"C:\\Users\\{os.getlogin()}\\Desktop\\{self.save_file_name}.xlsx")


if __name__ == '__main__':
    args = sys.argv[1]
    ActPost(f'{args}').create_document()
